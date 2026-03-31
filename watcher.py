import time
import os
import threading
import warnings

# Suppress requests/urllib3/chardet version mismatch warning (cosmetic)
warnings.filterwarnings("ignore", message=".*doesn't match a supported version.*")

import requests
import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from dotenv import load_dotenv

load_dotenv()

# ─── CONFIG ───────────────────────────────────────────────
FILE_TO_WATCH = os.getenv("file_to_watch")   # full path to Excel file
WATCH_DIR     = os.getenv("watch_dir")       # folder containing the file
N8N_WEBHOOK   = os.getenv("n8n_webhook")     # n8n webhook URL
DEBOUNCE_SEC  = int(os.getenv("watcher_debounce_sec", "12"))  # ignore rapid re-triggers
# ──────────────────────────────────────────────────────────


def get_row_count(filepath):
    """Returns the number of rows that actually have data (ignores deleted/empty rows)."""
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        return sum(
            1 for row in ws.iter_rows()
            if any(cell.value is not None for cell in row)
        )
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return None


def wait_until_file_stable(filepath, timeout=15):
    """Waits until the file size stops changing, meaning Excel has finished writing."""
    prev_size = -1
    elapsed   = 0
    while elapsed < timeout:
        time.sleep(1)
        elapsed += 1
        try:
            curr_size = os.path.getsize(filepath)
        except Exception:
            continue
        if curr_size == prev_size and curr_size > 0:
            return True  # file is stable
        prev_size = curr_size
    return False  # timed out


# Common Excel header names → API keys (case-insensitive match)
COMPANY_HEADERS = ("company_name", "company", "organization", "org", "name")
DOMAIN_HEADERS = ("domain", "website", "url", "web", "site")


def _normalize_lead(lead_dict):
    """Map Excel columns to API keys company_name and domain so n8n can forward as-is."""
    if not lead_dict:
        return {}
    out = {}
    keys_lower = {str(k).strip().lower(): k for k in lead_dict}
    for h in COMPANY_HEADERS:
        if h in keys_lower:
            out["company_name"] = lead_dict.get(keys_lower[h]) or ""
            break
    for h in DOMAIN_HEADERS:
        if h in keys_lower:
            out["domain"] = lead_dict.get(keys_lower[h]) or ""
            break
    return out if out.get("company_name") and out.get("domain") else {}


def get_last_row(filepath):
    """Returns the last row with actual data as a dict with column headers as keys."""
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        data_rows = [
            row for row in ws.iter_rows(values_only=True)
            if any(cell is not None for cell in row)
        ]

        if len(data_rows) < 2:
            return {}

        headers  = list(data_rows[0])   # first row = headers
        last_row = list(data_rows[-1])  # last row = newest lead
        raw = dict(zip(headers, last_row))
        # Send only API-shaped fields when we can map headers (avoids duplicate keys and huge payloads).
        normalized = _normalize_lead(raw)
        return normalized if normalized else raw

    except Exception as e:
        print(f"Error reading last row: {e}")
        return {}


class ExcelChangeHandler(FileSystemEventHandler):

    def __init__(self):
        self.last_row_count = get_row_count(FILE_TO_WATCH)
        self._lock = threading.Lock()
        self._last_processed_at = 0.0
        print(f"👀 Watching: {FILE_TO_WATCH}")
        print(f"📊 Current row count: {self.last_row_count}")
        print(f"⏱️  Debounce: {DEBOUNCE_SEC}s (set watcher_debounce_sec in .env to change)")

    def on_modified(self, event):
        if not event.src_path.endswith("Org_Domain.xlsx"):
            return

        with self._lock:
            if time.time() - self._last_processed_at < DEBOUNCE_SEC:
                print("⏭️  Skipping (debounce — too soon after last trigger)")
                return

        print("⏳ Waiting for Excel to finish writing...")
        wait_until_file_stable(FILE_TO_WATCH)

        with self._lock:
            new_row_count = get_row_count(FILE_TO_WATCH)
            if new_row_count is None:
                return

            print(f"🔍 Debug — last: {self.last_row_count}, current: {new_row_count}")

            if new_row_count > self.last_row_count:
                print(f"🆕 New row detected! {self.last_row_count} → {new_row_count}")

                lead_data = get_last_row(FILE_TO_WATCH)

                if not lead_data.get("company_name") or not lead_data.get("domain"):
                    print(
                        "⚠️  Tip: Use Excel headers like Organization/Domain (or Company/Website) "
                        "so the webhook sends only {\"company_name\",\"domain\"}. "
                        "In n8n, map $json.body.lead.company_name and $json.body.lead.domain."
                    )

                try:
                    response = requests.post(N8N_WEBHOOK, json={
                        "event": "new_lead",
                        "lead": lead_data
                    })
                    print(f"✅ n8n triggered! Status: {response.status_code}")
                    self.last_row_count = new_row_count
                    self._last_processed_at = time.time()
                except Exception as e:
                    print(f"❌ Failed to ping n8n: {e}")

            else:
                print("📝 File modified but no new row added (possible cell edit).")


if __name__ == "__main__":
    handler  = ExcelChangeHandler()
    observer = Observer()
    observer.schedule(handler, path=WATCH_DIR, recursive=False)
    observer.start()

    print("🚀 Watcher running... Press Ctrl+C to stop.\n")

    try:
        while True:
            time.sleep(2)
    except KeyboardInterrupt:
        observer.stop()
        print("\n🛑 Watcher stopped.")

    observer.join()